import uvicorn
from fastapi import FastAPI, UploadFile, File
from fastapi.responses import JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

app = FastAPI(title="VMI Leonisa API")

DB_PATH = os.path.join(os.path.dirname(__file__), "vmi_leonisa.db")
STATIC_DIR = os.path.dirname(__file__)

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS trascon (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_fact TEXT,
            tipo TEXT,
            nombre_corto TEXT,
            ref TEXT,
            refext TEXT,
            color TEXT,
            talla TEXT,
            cant_facturada REAL
        )
    """)
    conn.commit()
    conn.close()

init_db()

# ─── TRASCON ENDPOINTS ──────────────────────────────────────

@app.get("/api/trascon")
def get_trascon(search: str = ""):
    conn = get_db()
    if search:
        q = f"%{search}%"
        rows = conn.execute("""
            SELECT * FROM trascon 
            WHERE fecha_fact LIKE ? OR tipo LIKE ? OR nombre_corto LIKE ? 
               OR ref LIKE ? OR refext LIKE ? OR color LIKE ? 
               OR talla LIKE ? OR CAST(cant_facturada AS TEXT) LIKE ?
            ORDER BY id
        """, [q]*8).fetchall()
    else:
        rows = conn.execute("SELECT * FROM trascon ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/api/upload/trascon")
async def upload_trascon(file: UploadFile = File(...)):
    if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        return JSONResponse({"error": "Formato no soportado. Use .xlsx"}, status_code=400)

    content = await file.read()
    tmp_path = os.path.join(os.path.dirname(__file__), "_tmp_upload.xlsx")
    with open(tmp_path, "wb") as f:
        f.write(content)

    wb = load_workbook(tmp_path, read_only=True, data_only=True)
    ws = wb.active

    # Find headers and tipo column
    headers = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    tipo_idx = -1
    for i, h in enumerate(headers):
        if h.lower().replace(" ", "").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u") == "tipo":
            tipo_idx = i
            break

    if tipo_idx == -1:
        wb.close()
        os.remove(tmp_path)
        return JSONResponse({"error": "No se encontró la columna 'tipo'"}, status_code=400)

    conn = get_db()
    conn.execute("DELETE FROM trascon")
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= tipo_idx:
            continue
        tipo_val = str(row[tipo_idx] or "").lower()
        if "trascon" in tipo_val:
            fecha_fact = str(row[0] or "") if len(row) > 0 else ""
            tipo = str(row[1] or "") if len(row) > 1 else ""
            nombre = str(row[2] or "") if len(row) > 2 else ""
            ref = str(row[3] or "") if len(row) > 3 else ""
            refext = str(row[4] or "") if len(row) > 4 else ""
            color = str(row[5] or "") if len(row) > 5 else ""
            talla = str(row[6] or "") if len(row) > 6 else ""
            cant = 0
            if len(row) > 7 and row[7] is not None:
                try:
                    cant = float(row[7])
                except:
                    cant = 0
            conn.execute("""
                INSERT INTO trascon (fecha_fact, tipo, nombre_corto, ref, refext, color, talla, cant_facturada)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (fecha_fact, tipo, nombre, ref, refext, color, talla, cant))
            count += 1

    conn.commit()
    conn.close()
    wb.close()
    os.remove(tmp_path)
    return {"ok": True, "count": count}

@app.delete("/api/trascon")
def clear_trascon():
    conn = get_db()
    conn.execute("DELETE FROM trascon")
    conn.commit()
    conn.close()
    return {"ok": True}

@app.get("/api/trascon/count")
def trascon_count():
    conn = get_db()
    c = conn.execute("SELECT COUNT(*) FROM trascon").fetchone()[0]
    conn.close()
    return {"count": c}

# ─── SERVE STATIC FILES ─────────────────────────────────────

@app.get("/")
def serve_index():
    return FileResponse(os.path.join(STATIC_DIR, "index.html"))

@app.get("/{path:path}")
def serve_static(path: str):
    fp = os.path.join(STATIC_DIR, path)
    if os.path.isfile(fp):
        return FileResponse(fp)
    return JSONResponse({"error": "Not found"}, status_code=404)

if __name__ == "__main__":
    print("=" * 50)
    print(" VMI Leonisa API — http://localhost:8000")
    print("=" * 50)
    uvicorn.run(app, host="0.0.0.0", port=8000)
